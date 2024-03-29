import openpyxl

import json
import os


BASE_DIR = os.path.dirname(
    os.path.dirname(
        os.path.abspath(__file__)
    )
)

CSV_FILE = os.path.join(BASE_DIR, 'code.csv')
PREFIXES_FILE = os.path.join(BASE_DIR, 'prefix.json')
CONFIG_FILE = os.path.join(BASE_DIR, 'config.json')
UI_FOLDER = os.path.join(BASE_DIR, 'src', 'ui')
ADD_PREFIX_UI = os.path.join(UI_FOLDER, 'add-prefix.ui')
COLUMN_TITLE = 'Phone Numbers'
EXCEL_SUFFIX = 'xlsx'
docs_folder = 'docs'
star_folder = 'star'
number_folder = 'number'


class JsonConfigHandler:
    def __init__(self, file_path=CONFIG_FILE):
        self.file_path = file_path
        self.config = {}

    def load_config(self):
        try:
            with open(self.file_path, 'r') as file:
                self.config = json.load(file)
        except FileNotFoundError:
            return self.config

    def save_config(self):
        with open(self.file_path, 'w') as file:
            json.dump(self.config, file, indent=4)

    def set_config(self, key, value):
        self.config[key] = value

    def delete_config(self, key):
        self.config.pop(key, None)

    def get_config(self, key, default=None):
        return self.config.get(key, default)


class ExcelHandler:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.sheet = None

    def create_workbook(self):
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active

    def open_workbook(self):
        self.workbook = openpyxl.load_workbook(self.file_path)
        self.sheet = self.workbook.active

    def save_workbook(self):
        self.workbook.save(self.file_path)

    def close_workbook(self):
        self.workbook.close()

    def set_column_title(self, title):
        self.sheet.cell(row=1, column=1, value=title)

    def write_data(self, value):
        self.sheet.cell(row=self.sheet.max_row + 1, column=1, value=value)
